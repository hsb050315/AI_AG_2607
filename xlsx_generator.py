"""openpyxl 기반 범용 Excel 문서 생성 템플릿."""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_FONT = Font(name="맑은 고딕", size=18, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
DEFAULT_MAIN_COLOR = "1F4E79"  # 남색 계열

THIN = Side(style="thin", color="B7B7B7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def new_workbook(main_color=None):
    """새 Workbook 생성. main_color(hex, 예: '2E86AB')로 강조색을 지정한다."""
    wb = Workbook()
    wb.main_color = (main_color or DEFAULT_MAIN_COLOR).lstrip("#").upper()
    ws = wb.active
    return wb, ws


def add_title(ws, text, cell="A1", span_cols=1):
    """시트 상단에 제목을 추가하고 병합/색칠한다."""
    wb = ws.parent
    ws[cell] = text
    ws[cell].font = TITLE_FONT
    ws[cell].fill = PatternFill("solid", fgColor=wb.main_color)
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    row = ws[cell].row
    col = ws[cell].column
    if span_cols > 1:
        ws.merge_cells(
            start_row=row, start_column=col, end_row=row, end_column=col + span_cols - 1
        )
    ws.row_dimensions[row].height = 28
    return ws[cell]


def add_table(ws, headers, rows, start_row=1, start_col=1):
    """헤더 + 행 데이터를 표 형태로 추가 (헤더는 main_color 배경)."""
    wb = ws.parent
    fill = PatternFill("solid", fgColor=wb.main_color)

    for j, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + j, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for i, row_data in enumerate(rows, start=1):
        for j, value in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = CELL_BORDER

    auto_fit_columns(ws, headers, rows, start_col=start_col)
    return start_row + len(rows)


def auto_fit_columns(ws, headers, rows, start_col=1):
    """헤더/데이터 길이에 맞춰 열 너비를 자동 조정한다."""
    for j, header in enumerate(headers):
        col_letter = get_column_letter(start_col + j)
        max_len = len(str(header))
        for row_data in rows:
            if j < len(row_data):
                max_len = max(max_len, len(str(row_data[j])))
        ws.column_dimensions[col_letter].width = max_len + 4


def add_note(ws, text, cell):
    """일반 메모/설명 텍스트 추가."""
    ws[cell] = text
    ws[cell].font = BODY_FONT
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def freeze_header(ws, row=2):
    """지정한 행 아래를 고정(스크롤 시 헤더 유지)."""
    ws.freeze_panes = f"A{row}"


def save(wb, path):
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(path)
    print(f"저장 완료: {path}")


def build_sample_workbook(output_path="sample_output.xlsx", main_color=None):
    """템플릿 기능을 확인할 수 있는 예시 엑셀 생성."""
    wb, ws = new_workbook(main_color=main_color)
    ws.title = "요약"

    add_title(ws, "샘플 데이터 시트", cell="A1", span_cols=4)

    headers = ["항목", "가격", "재고", "비고"]
    rows = [
        ["상품 A", 10000, 23, "-"],
        ["상품 B", 15000, 0, "품절"],
        ["상품 C", 8000, 120, "-"],
    ]
    last_row = add_table(ws, headers, rows, start_row=3, start_col=1)
    freeze_header(ws, row=4)

    add_note(ws, "이 시트는 xlsx_generator.py 기능 확인용 예시입니다.", cell=f"A{last_row + 2}")

    save(wb, output_path)


if __name__ == "__main__":
    build_sample_workbook(main_color="2E86AB")
